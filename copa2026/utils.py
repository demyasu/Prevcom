import math
from datetime import datetime, timedelta
from data import GROUPS_DATA, ALL_TEAMS, BRACKET_MATCHUPS

def calcular_probabilidade(time_casa, time_fora):
    ranking_casa = next((t["ranking"] for t in ALL_TEAMS if t["name"] == time_casa), 50)
    ranking_fora = next((t["ranking"] for t in ALL_TEAMS if t["name"] == time_fora), 50)

    diff = ranking_fora - ranking_casa
    expected_casa = 1.0 / (1.0 + 10 ** (-diff / 40.0))

    ranking_gap = abs(ranking_casa - ranking_fora)
    prob_empate = max(0.08, 0.35 - ranking_gap * 0.004)

    prob_casa = expected_casa * (1.0 - prob_empate)
    prob_fora = (1.0 - expected_casa) * (1.0 - prob_empate)

    total = prob_casa + prob_empate + prob_fora
    prob_casa /= total
    prob_empate /= total
    prob_fora /= total

    media_gols = max(1.2, 3.0 - ranking_gap / 50.0)
    odds_casa = round(1.0 / max(0.01, prob_casa), 2)
    odds_empate = round(1.0 / max(0.01, prob_empate), 2)
    odds_fora = round(1.0 / max(0.01, prob_fora), 2)

    return {
        "casa": {
            "nome": time_casa,
            "probabilidade": round(prob_casa * 100, 1),
            "odds": odds_casa,
        },
        "empate": {
            "probabilidade": round(prob_empate * 100, 1),
            "odds": odds_empate,
        },
        "fora": {
            "nome": time_fora,
            "probabilidade": round(prob_fora * 100, 1),
            "odds": odds_fora,
        },
        "media_gols_esperados": round(media_gols, 2),
        "mercados": {
            "ambos_marcam_sim": round(0.55 + (media_gols - 0.5) * 0.1, 2),
            "ambos_marcam_nao": round(0.45 - (media_gols - 0.5) * 0.1, 2),
            "mais_25_gols": round(0.45 + (media_gols - 0.5) * 0.15, 2),
            "menos_25_gols": round(0.55 - (media_gols - 0.5) * 0.15, 2),
        }
    }

def exportar_xlsx(filepath="copa2026_chaveamento.xlsx", standings=None, match_scores=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from data import MATCH_DATES, SUFFIX_LEGEND, GROUP_MATCHES, dia_semana

    if standings is None:
        standings = {}
    if match_scores is None:
        match_scores = {}

    wb = Workbook()
    header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    cell_font = Font(size=11)
    title_font = Font(bold=True, size=13, color="1B5E20")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')

    # Sheet 1: Grupos
    ws_grupos = wb.active
    ws_grupos.title = "Grupos"

    row = 1
    for g, data in GROUPS_DATA.items():
        ws_grupos.cell(row=row, column=1, value=f"Grupo {g}").font = title_font
        row += 1
        headers = ["#", "Seleção", "Pts", "J", "V", "E", "D", "GP", "GC", "SG", "Ranking", "Confederação"]
        for col, h in enumerate(headers, 1):
            cell = ws_grupos.cell(row=row, column=col, value=h)
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = thin_border
        row += 1
        group_standings = standings.get(g, [])
        for i, t in enumerate(data["teams"], 1):
            st = next((s for s in group_standings if s["name"] == t["name"]), {})
            vals = [
                i, t["name"],
                st.get("points", 0), st.get("played", 0),
                st.get("wins", 0), st.get("draws", 0), st.get("losses", 0),
                st.get("gf", 0), st.get("ga", 0), st.get("gd", 0),
                t["ranking"], t["confederation"]
            ]
            for col, val in enumerate(vals, 1):
                cell = ws_grupos.cell(row=row, column=col, value=val)
                cell.font = cell_font; cell.border = thin_border
            row += 1

        matches = MATCH_DATES["group"].get(g, [])
        if matches:
            row += 1
            ws_grupos.cell(row=row, column=1, value="Jogos:").font = Font(bold=True, size=11)
            row += 1
            for col, h in enumerate(["Data", "Horário", "Casa", "Placar", "Fora", "Estádio"], 1):
                cell = ws_grupos.cell(row=row, column=col, value=h)
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = center; cell.border = thin_border
            row += 1
            group_match_ids = {m["id"]: m for m in GROUP_MATCHES.get(g, [])}
            for m in matches:
                mid = next((k for k, v in group_match_ids.items() if v["home"] == m["home"] and v["away"] == m["away"]), None)
                placar = ""
                if mid and mid in match_scores:
                    ms = match_scores[mid]
                    placar = f"{ms['home_score']} x {ms['away_score']}"
                dw = dia_semana(m["date"])
                data_label = f"{dw} {m['date']}" if dw else m["date"]
                for col, val in enumerate([data_label, m["time"], m["home"], placar, m["away"], m["venue"]], 1):
                    cell = ws_grupos.cell(row=row, column=col, value=val)
                    cell.font = cell_font; cell.border = thin_border
                row += 1
        row += 1

    # Sheet 2: Probabilidades
    ws_prob = wb.create_sheet("Probabilidades")
    prob_headers = ["Fase", "Partida", "Data", "Time Casa", "Casa %", "Casa Odds", "Empate %", "Empate Odds", "Time Fora", "Fora %", "Fora Odds", "Gols Esperados"]
    for col, h in enumerate(prob_headers, 1):
        cell = ws_prob.cell(row=1, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = thin_border

    STAGE_LABELS = {
        "round_of_32": "16-avos", "round_of_16": "Oitavas", "quarter_finals": "Quartas",
        "semi_finals": "Semifinais", "third_place": "3º Lugar", "final": "Final"
    }
    row = 2
    for stage_key, stage in BRACKET_MATCHUPS.items():
        stage_dates = MATCH_DATES.get("knockout", {}).get(stage_key, {})
        for m in stage:
            md = stage_dates.get(m["id"], {})
            home_label = m["home"]; away_label = m["away"]
            prob = calcular_probabilidade(home_label, away_label)
            dt = md.get("date","")
            dw = dia_semana(dt) if dt else ""
            data_label = f"{dw} {dt} {md.get('time','')}" if dw else f"{dt} {md.get('time','')}".strip()
            vals = [
                STAGE_LABELS.get(stage_key, stage_key), m["id"],
                data_label,
                prob["casa"]["nome"], f"{prob['casa']['probabilidade']}%", prob["casa"]["odds"],
                f"{prob['empate']['probabilidade']}%", prob["empate"]["odds"],
                prob["fora"]["nome"], f"{prob['fora']['probabilidade']}%", prob["fora"]["odds"],
                prob["media_gols_esperados"]
            ]
            for col, val in enumerate(vals, 1):
                cell = ws_prob.cell(row=row, column=col, value=val)
                cell.font = cell_font; cell.border = thin_border
            row += 1

    # Sheet 3: Chaveamento
    ws_bracket = wb.create_sheet("Chaveamento")
    row = 1
    for stage_key, stage_label in [
        ("round_of_32", "16-avos de Final (28 Jun - 3 Jul)"),
        ("round_of_16", "Oitavas de Final (4-7 Jul)"),
        ("quarter_finals", "Quartas de Final (9-11 Jul)"),
        ("semi_finals", "Semifinais (14-15 Jul)"),
        ("third_place", "Disputa de 3º Lugar (18 Jul)"),
        ("final", "Final (19 Jul)")
    ]:
        stage_data = BRACKET_MATCHUPS.get(stage_key, [])
        if not stage_data:
            continue
        ws_bracket.cell(row=row, column=1, value=stage_label).font = title_font
        row += 1
        headers = ["Partida", "Data", "Horário", "Casa", "Placar", "Fora", "Estádio", "Avança para"]
        for col, h in enumerate(headers, 1):
            cell = ws_bracket.cell(row=row, column=col, value=h)
            cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = thin_border
        row += 1
        stage_dates = MATCH_DATES.get("knockout", {}).get(stage_key, {})
        for m in stage_data:
            md = stage_dates.get(m["id"], {})
            dt = md.get("date","")
            dw = dia_semana(dt) if dt else ""
            data_label = f"{dw} {dt}" if dw else dt
            vals = [m["id"], data_label, md.get("time",""), m["home"], "", m["away"], md.get("venue",""), m.get("next","")]
            for col, val in enumerate(vals, 1):
                cell = ws_bracket.cell(row=row, column=col, value=val)
                cell.font = cell_font; cell.border = thin_border
            row += 1
        row += 1

    # Sheet 4: Todas as Seleções
    ws_all = wb.create_sheet("Todas as Seleções")
    all_headers = ["Seleção", "Grupo", "Ranking FIFA", "Confederação"]
    for col, h in enumerate(all_headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = thin_border
    row = 2
    sorted_teams = sorted(ALL_TEAMS, key=lambda x: x["ranking"])
    for t in sorted_teams:
        for col, val in enumerate([t["name"], t["group"], t["ranking"], t["confederation"]], 1):
            cell = ws_all.cell(row=row, column=col, value=val)
            cell.font = cell_font; cell.border = thin_border
        row += 1

    # Sheet 5: Legenda
    ws_leg = wb.create_sheet("Legenda")
    ws_leg.cell(row=1, column=1, value="Legenda do Chaveamento").font = title_font
    for col, h in enumerate(["Sigla", "Significado"], 1):
        cell = ws_leg.cell(row=2, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font; cell.alignment = center; cell.border = thin_border
    row = 3
    for sigla, desc in SUFFIX_LEGEND.items():
        for col, val in enumerate([sigla, desc], 1):
            cell = ws_leg.cell(row=row, column=col, value=val)
            cell.font = cell_font; cell.border = thin_border
        row += 1

    # Ajustar largura das colunas
    for ws in [ws_grupos, ws_prob, ws_bracket, ws_all, ws_leg]:
        for col_cells in ws.columns:
            max_length = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 3, 40)

    wb.save(filepath)
    return filepath
