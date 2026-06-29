from flask import Flask, render_template, request, jsonify, session, Response
import io
from models import db, Questao, Flashcard, SimuladoResultado, Resposta
from datetime import datetime, date
import random
import json
import os
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'fallback-dev-key')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///prevcom.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

PROVA_DATE = date(2026, 9, 13)

# No cache
@app.after_request
def no_cache(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.context_processor
def inject_globals():
    dias_restantes = (PROVA_DATE - date.today()).days
    return dict(dias_restantes=dias_restantes)

with app.app_context():
    db.create_all()
    from seed_data import seed_all
    seed_all(app)

# ====== HOME ======
@app.route('/')
def index():
    total_questoes = Questao.query.count()
    total_flashcards = Flashcard.query.count()
    simulados_feitos = SimuladoResultado.query.count()
    stats = db.session.query(db.func.avg(SimuladoResultado.acertos), db.func.max(SimuladoResultado.acertos)).first()
    desempenho = {
        'media': round(stats[0], 1) if stats[0] else 0,
        'maximo': stats[1] or 0,
        'total_questoes': total_questoes,
        'total_flashcards': total_flashcards,
        'simulados_feitos': simulados_feitos
    }
    return render_template('index.html', desempenho=desempenho)

# ====== SIMULADO ======
@app.route('/simulado')
def simulado():
    return render_template('simulado.html')

@app.route('/api/simulado/gerar', methods=['POST'])
def gerar_simulado():
    data = request.get_json() or {}
    num = data.get('quantidade', 10)
    discs = data.get('disciplinas', [])
    q = Questao.query
    if discs:
        q = q.filter(Questao.disciplina.in_(discs))
    total = q.count()
    if total == 0:
        return jsonify({'erro': 'Nenhuma questão disponível'}), 404
    questoes = q.order_by(db.func.random()).limit(min(num, total)).all()
    session['simulado_atual'] = {
        'questoes_ids': [q.id for q in questoes],
        'respostas': {},
        'inicio': datetime.now().isoformat()
    }
    return jsonify([{
        'id': q.id, 'disciplina': q.disciplina, 'enunciado': q.enunciado,
        'alternativas': {'a': q.alternativa_a, 'b': q.alternativa_b, 'c': q.alternativa_c, 'd': q.alternativa_d, 'e': q.alternativa_e}
    } for q in questoes])

@app.route('/api/simulado/responder', methods=['POST'])
def responder_questao():
    data = request.get_json() or {}
    qid = data.get('questao_id')
    alt = data.get('alternativa')
    if 'simulado_atual' not in session:
        return jsonify({'erro': 'Nenhum simulado em andamento'}), 400
    session['simulado_atual']['respostas'][str(qid)] = alt
    session.modified = True
    q = Questao.query.get(qid)
    return jsonify({'correta': q.alternativa_correta == alt, 'gabarito': q.alternativa_correta, 'explicacao': q.explicacao})

@app.route('/api/simulado/finalizar', methods=['POST'])
def finalizar_simulado():
    if 'simulado_atual' not in session:
        return jsonify({'erro': 'Nenhum simulado em andamento'}), 400
    sd = session['simulado_atual']
    acertos = 0
    detalhes = []
    for qid in sd['questoes_ids']:
        q = Questao.query.get(qid)
        esc = sd['respostas'].get(str(qid), '')
        corr = esc == q.alternativa_correta
        if corr: acertos += 1
        detalhes.append({'questao_id': qid, 'enunciado': q.enunciado, 'sua_resposta': esc, 'gabarito': q.alternativa_correta, 'correta': corr, 'explicacao': q.explicacao})
    res = SimuladoResultado(total=len(sd['questoes_ids']), acertos=acertos, data=datetime.now())
    db.session.add(res)
    db.session.flush()
    for qid in sd['questoes_ids']:
        q = Questao.query.get(qid)
        esc = sd['respostas'].get(str(qid), '')
        corr = esc == q.alternativa_correta
        db.session.add(Resposta(questao_id=qid, simulado_id=res.id, alternativa_escolhida=esc or None, correta=corr, errou=not corr))
    db.session.commit()
    session.pop('simulado_atual', None)
    return jsonify({'total': len(sd['questoes_ids']), 'acertos': acertos, 'percentual': round(acertos/len(sd['questoes_ids'])*100, 1), 'detalhes': detalhes})

@app.route('/api/disciplinas')
def listar_disciplinas():
    return jsonify([d[0] for d in db.session.query(Questao.disciplina).distinct().all()])

# ====== FLASHCARDS ======
@app.route('/flashcards')
def flashcards():
    return render_template('flashcards.html')

@app.route('/api/flashcards')
def listar_flashcards():
    disc = request.args.get('disciplina')
    q = Flashcard.query
    if disc:
        q = q.filter(Flashcard.disciplina == disc)
    cards = q.order_by(Flashcard.erros.desc(), Flashcard.revisoes.asc()).all()
    return jsonify([c.to_dict() for c in cards])

@app.route('/api/flashcards/revisar', methods=['POST'])
def revisar_flashcard():
    data = request.get_json() or {}
    card = Flashcard.query.get(data.get('id'))
    if not card:
        return jsonify({'erro': 'Flashcard não encontrado'}), 404
    if data.get('acertou'):
        card.acertos += 1
    else:
        card.erros += 1
    card.revisoes += 1
    db.session.commit()
    return jsonify({'status': 'ok', 'card': card.to_dict()})

@app.route('/api/flashcards/adicionar', methods=['POST'])
def adicionar_flashcard():
    data = request.get_json() or {}
    card = Flashcard(frente=data['frente'], verso=data['verso'], disciplina=data.get('disciplina', 'Geral'))
    db.session.add(card)
    db.session.commit()
    return jsonify({'id': card.id, 'status': 'ok'})

@app.route('/api/flashcards/contagem')
def contagem_flashcards():
    from collections import defaultdict
    counts = defaultdict(int)
    for disc, count in db.session.query(Flashcard.disciplina, db.func.count(Flashcard.id)).group_by(Flashcard.disciplina).all():
        counts[disc] = count
    return jsonify(dict(counts))

# ====== EXPORT ======
@app.route('/api/flashcards/exportar/xlsx')
def exportar_flashcards_xlsx():
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flashcards PREVCOM"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    ws.append(["Disciplina", "Frente (Conceito)", "Verso (Definição)", "Acertos", "Erros", "Revisões", "Dificuldade"])
    for col in range(1, 8):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 80
    for card in Flashcard.query.order_by(Flashcard.disciplina, Flashcard.id).all():
        ws.append([card.disciplina, card.frente, card.verso, card.acertos, card.erros, card.revisoes, f"{card.dificuldade}/4"])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=flashcards_prevcom.xlsx"})

@app.route('/api/flashcards/exportar/pdf')
def exportar_flashcards_pdf():
    from fpdf import FPDF
    _dir = os.path.dirname(__file__)
    pdf = FPDF()
    pdf.add_font("DejaVu", "", os.path.join(_dir, "fonts", "DejaVuSans.ttf"))
    pdf.add_font("DejaVu", "B", os.path.join(_dir, "fonts", "DejaVuSans-Bold.ttf"))
    pdf.add_font("DejaVu", "I", os.path.join(_dir, "fonts", "DejaVuSans-Oblique.ttf"))
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("DejaVu", "B", 18)
    pdf.set_text_color(13, 110, 253)
    pdf.cell(0, 12, "PREVCOM 2026 - Flashcards", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("DejaVu", "I", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, "Total de cards: {}".format(Flashcard.query.count()), align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    disc_atual = ""
    for card in Flashcard.query.order_by(Flashcard.disciplina, Flashcard.id).all():
        if pdf.get_y() > 250:
            pdf.add_page()
        if card.disciplina != disc_atual:
            disc_atual = card.disciplina
            pdf.ln(2)
            pdf.set_font("DejaVu", "B", 13)
            pdf.set_text_color(13, 110, 253)
            pdf.cell(0, 8, disc_atual, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(1)
        pdf.set_font("DejaVu", "B", 9)
        pdf.set_text_color(30, 30, 30)
        pdf.multi_cell(0, 4.5, "Q: " + card.frente, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("DejaVu", "", 9)
        pdf.set_text_color(60, 60, 60)
        pdf.multi_cell(0, 4.5, "R: " + card.verso, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("DejaVu", "I", 7)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 3, "Dif: {}/4 | Acertos: {} | Erros: {}".format(card.dificuldade, card.acertos, card.erros), new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1.5)
    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return Response(output.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": "attachment;filename=flashcards_prevcom.pdf"})

# ====== CONCEITOS EXPORT ======
@app.route('/api/conceitos/exportar/pdf')
def exportar_conceitos_pdf():
    from fpdf import FPDF
    _dir = os.path.dirname(__file__)
    pdf = FPDF()
    pdf.add_font("DejaVu", "", os.path.join(_dir, "fonts", "DejaVuSans.ttf"))
    pdf.add_font("DejaVu", "B", os.path.join(_dir, "fonts", "DejaVuSans-Bold.ttf"))
    pdf.add_font("DejaVu", "I", os.path.join(_dir, "fonts", "DejaVuSans-Oblique.ttf"))
    pdf.add_page()
    pdf.set_font("DejaVu", "B", 18)
    pdf.set_text_color(13, 110, 253)
    pdf.cell(0, 12, "PREVCOM 2026 - Conceitos Detalhados", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("DejaVu", "I", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, "Material completo do edital", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    secoes = [
        ("Lingua Portuguesa", [
            ("1. Leitura e Interpretacao", "Compreensao = extrair informacoes explicitas. Interpretacao = inferir informacoes implicitas."),
            ("2. Tipologia Textual", "Narracao (conta historia), Descricao (caracteriza), Dissertacao (defende ideias), Injuncao (instrui)."),
            ("3. Figuras de Linguagem", "Metafora, Metonimia, Comparacao, Antitese, Paradoxo, Hiperbole, Eufemismo, Ironia, Prosopopeia."),
            ("4. Concordancia Verbal", "Verbo concorda com sujeito em numero e pessoa. Casos: Haver impessoal (sempre singular), Fazer temporal (singular), Maioria (singular/plural)."),
            ("5. Concordancia Nominal", "Adjetivo concorda com substantivo. Casos: E proibido (invariável sem artigo), Anexo (concorda), Menos (invariável)."),
            ("6. Regencia Verbal", "Assistir (ver = a), Aspirar (desejar = a), Visar (objetivar = a), Implicar (direto), Obedecer (a)."),
            ("7. Colocacao Pronominal", "Proclise (antes: negacao, relativos, adverbios), Enclise (depois: inicio, imperativo), Mesoclise (futuro)."),
            ("8. Crase", "Fusao a+a. Obrigatoria: palavras femininas, locucoes, aquele. Proibida: masculinos, verbos, pronomes."),
        ]),
        ("Etica e Integridade", [
            ("1. Principios LIMPE", "Legalidade, Impessoalidade, Moralidade, Publicidade, Eficiencia (art. 37 CF)."),
            ("2. Programa de Integridade", "Prevencao, deteccao e correcao de desvios. Pilares: comprometimento, codigo, treinamento, canal, investigacao."),
            ("3. ESG", "Environmental (clima, residuos), Social (direitos humanos, diversidade), Governance (transparencia, compliance)."),
            ("4. Lei 8.429/92 (Improbidade)", "Enriquecimento ilicito (art. 9), Lesao ao erario (art. 10), Atentado aos principios (art. 11). Exige dolo."),
            ("5. Lei 12.846/13 (Anticorrupcao)", "Responsabilidade objetiva de PJ. Multa 0,1%-30% do faturamento. Acordo de leniencia."),
            ("6. Decreto 1.171/94", "Codigo de Etica do Servidor. Termo de compromisso na posse. Censura etica."),
            ("7. LGPD", "Lei 13.709/18. Dados pessoais. Controlador, operador, titular. ANPD. Multa ate 2% (limite R$50M)."),
        ]),
        ("Raciocinio Logico", [
            ("1. Proposicoes", "Declarativas V/F. Operadores: ~, ^, v, ->, <->. Tabela verdade."),
            ("2. Equivalencias", "p->q = ~p v q = ~q->~p. De Morgan: ~(p^q) = ~p v ~q."),
            ("3. Quantificadores", "Todo (universal), Algum (existencial). Negacao: troca quantificador e nega predicado."),
            ("4. Argumentacao", "Modus Ponens, Modus Tollens, Silogismo. Validado x falacia."),
            ("5. Conjuntos", "Uniao, intersecao, diferenca. Formulas n(AUB) = n(A)+n(B)-n(A interseccao B)."),
            ("6. Probabilidade", "P(E) = favoraveis/possiveis. Independentes: P(A interseccao B)=P(A)*P(B). Bayes: P(A|B)=P(B|A)*P(A)/P(B)."),
            ("7. Sequencias", "PA: an=a1+(n-1)r. PG: an=a1*q^(n-1). Fibonacci: cada termo = soma dos 2 anteriores."),
        ]),
        ("Previdencia Complementar", [
            ("1. Regimes", "RGPS (INSS), RPPS (servidores), RPC (facultativo). Capitalizacao vs Reparticao."),
            ("2. Orgaos", "CNPC (normativo), PREVIC (fiscaliza EFPC), SUSEP (fiscaliza EAPC), CMN (investimentos)."),
            ("3. EFPC vs EAPC", "EFPC: fechada, sem fins lucrativos, PREVIC. EAPC: aberta, bancos, SUSEP."),
            ("4. LC 108/2001", "Patrocinio publico. Paridade contribuicao. Segregacao de massas."),
            ("5. LC 109/2001", "Lei base do RPC. Planos BD, CD, CV. Portabilidade, Resgate, BPD, Autopatrocínio."),
            ("6. PREVCOM SP", "Lei 14.653/2011. Decreto 57.785/2012. Estrutura: CD, CF, DIREX."),
            ("7. Tributacao", "PGBL (dedutivel ate 12%, IR no total). VGBL (nao dedutivel, IR so rendimentos). Tabela progressiva x regressiva."),
        ]),
        ("Conhecimentos Especificos", [
            ("1. Engenharia de Software", "Cascata, Scrum (papeis, cerimonias, artefatos), Kanban, XP. UML: casos de uso, classes, sequencia."),
            ("2. Banco de Dados", "SQL (SELECT, JOIN, GROUP BY), Normalizacao (1FN, 2FN, 3FN), NoSQL (MongoDB, Redis)."),
            ("3. Redes", "TCP/IP (4 camadas), OSI (7 camadas). HTTP/HTTPS, DNS, DHCP. Firewall, VPN."),
            ("4. Seguranca", "OWASP Top 10, Criptografia (simetrica/assimetrica), LGPD, ISO 27001."),
            ("5. Cloud", "IaaS, PaaS, SaaS. AWS, Azure, GCP. Docker, Kubernetes. DevOps, CI/CD."),
            ("6. Design Patterns", "GoF: Singleton, Factory, Strategy, Observer, Adapter. SOLID."),
            ("7. Dados", "Estatistica descritiva e inferencial. ML supervisionado e nao supervisionado. Python: Pandas, NumPy, Scikit-learn."),
        ]),
    ]
    for titulo, topicos in secoes:
        if pdf.get_y() > 240:
            pdf.add_page()
        pdf.set_font("DejaVu", "B", 14)
        pdf.set_text_color(13, 110, 253)
        pdf.cell(0, 9, titulo, new_x="LMARGIN", new_y="NEXT")
        for topico, desc in topicos:
            pdf.set_font("DejaVu", "B", 10)
            pdf.set_text_color(50, 50, 50)
            pdf.cell(0, 6, topico, new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("DejaVu", "", 9)
            pdf.set_text_color(80, 80, 80)
            pdf.multi_cell(0, 4.5, desc)
            pdf.ln(1)
        pdf.ln(3)
    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return Response(output.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": "attachment;filename=conceitos_prevcom.pdf"})

# ====== CONCEITOS EXPORT INDIVIDUAL ======
MATERIAS = {
    "lp": "lingua_portuguesa",
    "etica": "etica_integridade",
    "rl": "raciocinio_logico",
    "prev": "previdencia_complementar",
    "especificos": "conhecimentos_especificos",
}

def gerar_pdf_conceitos(materia_id):
    from conceitos_data import CONCEITOS
    from fpdf import FPDF
    m = CONCEITOS[MATERIAS[materia_id]]
    _dir = os.path.dirname(__file__)
    pdf = FPDF()
    pdf.add_font("DejaVu", "", os.path.join(_dir, "fonts", "DejaVuSans.ttf"))
    pdf.add_font("DejaVu", "B", os.path.join(_dir, "fonts", "DejaVuSans-Bold.ttf"))
    pdf.add_font("DejaVu", "I", os.path.join(_dir, "fonts", "DejaVuSans-Oblique.ttf"))
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("DejaVu", "B", 18)
    pdf.set_text_color(13, 110, 253)
    pdf.cell(0, 12, m["titulo"], align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("DejaVu", "I", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, "PREVCOM 2026 - Conceitos Detalhados", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    for titulo, desc in m["topicos"]:
        if pdf.get_y() > 245:
            pdf.add_page()
        pdf.set_font("DejaVu", "B", 11)
        pdf.set_text_color(13, 110, 253)
        pdf.multi_cell(0, 5.5, titulo, new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("DejaVu", "", 9)
        pdf.set_text_color(60, 60, 60)
        pdf.multi_cell(0, 4.5, desc, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1.5)
    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    filename = materia_id + "_prevcom.pdf"
    return Response(output.getvalue(), mimetype="application/pdf",
                    headers={"Content-Disposition": f"attachment;filename={filename}"})

def gerar_xlsx_conceitos(materia_id):
    from conceitos_data import CONCEITOS
    import openpyxl
    from openpyxl.styles import Font, Alignment
    m = CONCEITOS[MATERIAS[materia_id]]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = m["titulo"][:31]
    ws.append([m["titulo"]])
    ws["A1"].font = Font(bold=True, size=14, color="0D6EFD")
    ws.merge_cells("A1:B1")
    ws.append(["Topico", "Descricao"])
    ws["A2"].font = Font(bold=True, size=11)
    ws["B2"].font = Font(bold=True, size="11")
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 90
    for titulo, desc in m["topicos"]:
        ws.append([titulo, desc])
        for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = materia_id + "_prevcom.xlsx"
    return Response(output.getvalue(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={filename}"})

for mid in MATERIAS:
    def _make_pdf(mid=mid):
        return lambda: gerar_pdf_conceitos(mid)
    def _make_xlsx(mid=mid):
        return lambda: gerar_xlsx_conceitos(mid)
    app.add_url_rule(f'/api/conceitos/exportar/{mid}/pdf', f'exportar_conceitos_{mid}_pdf', _make_pdf())
    app.add_url_rule(f'/api/conceitos/exportar/{mid}/xlsx', f'exportar_conceitos_{mid}_xlsx', _make_xlsx())

# ====== ESTATÍSTICAS ======
@app.route('/estatisticas')
def estatisticas():
    return render_template('estatisticas.html')

@app.route('/api/estatisticas')
def api_estatisticas():
    hist = []
    for r in SimuladoResultado.query.order_by(SimuladoResultado.data).all():
        hist.append({'data': r.data.strftime('%d/%m/%Y'), 'acertos': r.acertos, 'total': r.total, 'percentual': round(r.acertos/r.total*100, 1)})
    discs = [{'nome': d[0], 'total': d[1]} for d in db.session.query(Questao.disciplina, db.func.count(Questao.id)).group_by(Questao.disciplina).all()]
    fc = db.session.query(db.func.count(Flashcard.id), db.func.sum(Flashcard.acertos), db.func.sum(Flashcard.erros)).first()
    fc_disc = []
    for d in db.session.query(Flashcard.disciplina, db.func.count(Flashcard.id), db.func.sum(Flashcard.acertos), db.func.sum(Flashcard.erros), db.func.sum(Flashcard.revisoes)).group_by(Flashcard.disciplina).all():
        fc_disc.append({'nome': d[0], 'total': d[1], 'acertos': d[2] or 0, 'erros': d[3] or 0, 'revisoes': d[4] or 0})
    return jsonify({'historico': hist, 'disciplinas': discs, 'flashcards': {'total': fc[0] or 0, 'acertos': fc[1] or 0, 'erros': fc[2] or 0, 'por_disciplina': fc_disc}, 'dias_ate_prova': (PROVA_DATE - date.today()).days})

# ====== RECOMENDAÇÕES ======
@app.route('/api/recomendacoes')
def recomendacoes():
    RECS_PADRAO = [
        {'topico': 'Língua Portuguesa', 'motivo': 'Maior peso (20 questões). Base essencial.', 'prioridade': 'alta', 'tipo': 'padrao'},
        {'topico': 'Previdência Complementar', 'motivo': '20 questões. Leis 108/109.', 'prioridade': 'alta', 'tipo': 'padrao'},
        {'topico': 'Raciocínio Lógico', 'motivo': '15 questões. Estruturas lógicas.', 'prioridade': 'media', 'tipo': 'padrao'},
        {'topico': 'Conhecimentos Específicos', 'motivo': '40 questões. Eng. Software, Dados, Segurança.', 'prioridade': 'alta', 'tipo': 'padrao'},
        {'topico': 'Ética e Integridade', 'motivo': '5 questões. ESG, compliance.', 'prioridade': 'media', 'tipo': 'padrao'}
    ]
    recs = []
    # 1. Simulados
    if SimuladoResultado.query.count() > 0:
        for qid, _ in db.session.query(Resposta.questao_id, db.func.count(Resposta.errou)).filter(Resposta.errou==True).group_by(Resposta.questao_id).order_by(db.func.count(Resposta.errou).desc()).limit(5).all():
            q = Questao.query.get(qid)
            if q:
                recs.append({'topico': f'{q.disciplina}: {q.enunciado[:70]}...', 'motivo': 'Errado em simulado.', 'prioridade': 'alta', 'tipo': 'simulado'})
    # 2. Flashcards por disciplina
    for disc, erros, acertos in db.session.query(Flashcard.disciplina, db.func.sum(Flashcard.erros), db.func.sum(Flashcard.acertos)).group_by(Flashcard.disciplina).having(db.func.sum(Flashcard.erros)+db.func.sum(Flashcard.acertos)>0).all():
        tot = (erros or 0)+(acertos or 0)
        if tot and (erros or 0)/tot >= 0.3:
            recs.append({'topico': f'Flashcards: {disc}', 'motivo': f'{erros} erros em {tot} tentativas.', 'prioridade': 'alta' if (erros or 0)/tot >= 0.5 else 'media', 'tipo': 'flashcard'})
    # 3. Cards difíceis
    for c in Flashcard.query.filter(Flashcard.erros+Flashcard.acertos>0, Flashcard.erros>=Flashcard.acertos).order_by(Flashcard.erros.desc()).limit(3):
        recs.append({'topico': f'Card: {c.frente[:60]}', 'motivo': f'{c.erros}E/{c.acertos}A em {c.disciplina}.', 'prioridade': 'alta', 'tipo': 'flashcard'})
    if not recs:
        return jsonify(RECS_PADRAO)
    # dedup
    uniq, seen = [], set()
    for r in recs:
        k = r['topico'][:40]
        if k not in seen:
            seen.add(k)
            uniq.append(r)
    return jsonify(uniq[:8])

# ====== EDITAL / OUTROS ======
@app.route('/edital')
def edital():
    return render_template('edital.html')

@app.route('/api/edital')
def api_edital():
    with open(os.path.join(app.root_path, 'data', 'conteudo_programatico.json'), 'r', encoding='utf-8') as f:
        return jsonify(json.load(f))

@app.route('/provas-anteriores')
def provas_anteriores():
    return render_template('provas_anteriores.html')

@app.route('/conceitos')
def conceitos():
    return render_template('conceitos.html')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)